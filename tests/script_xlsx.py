from openpyxl import load_workbook

workbook = load_workbook('tmp/file_example_XLSX_50.xlsx')
sheet = workbook.active
cell_value = sheet.cell(row=1, column=2).value
print(cell_value)



